# quantflow/utils/io_utils.py
"""
QuantFlow Financial Suite - I/O Utilities
Comprehensive file handling, data export, and format conversion utilities
"""

import pandas as pd
import numpy as np
import json
import yaml
import pickle
from pathlib import Path
from typing import Dict, List, Optional, Union, Any, Tuple
import logging
from datetime import datetime, date
import warnings
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart import LineChart, Reference
import matplotlib.pyplot as plt
import seaborn as sns
from ..config import get_config

logger = logging.getLogger(__name__)

warnings.filterwarnings("ignore", category=FutureWarning)
pd.set_option('future.no_silent_downcasting', True)

class DataExporter:
    """
    Comprehensive data export utility for QuantFlow financial data
    Supports multiple formats with professional styling
    """
    
    def __init__(self, config=None):
        self.config = config or get_config()
        self.template_path = self.config.base_path / "templates"
        self.export_history = []
    
    def create_dcf_report(self, 
                         ticker: str,
                         dcf_results,
                         financial_data: Dict,
                         output_format: str = 'html') -> str:
        """
        Create comprehensive DCF analysis report
        
        Args:   
            ticker: Stock ticker
            dcf_results: DCF calculation results
            financial_data: Complete financial data
            output_format: Output format ('html', 'pdf', 'excel')
            
        Returns:
            Path to created report
        """
        logger.info(f"Creating DCF report for {ticker}")
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if output_format == 'excel':
            filename = f"{ticker}_DCF_Report_{timestamp}.xlsx"
            return self.export_dcf_model_to_excel(dcf_results, financial_data, filename)
        
        elif output_format == 'html':
            return self._create_html_dcf_report(ticker, dcf_results, financial_data, timestamp)
        
        else:
            raise ValueError(f"Unsupported output format: {output_format}")
    
    def _create_html_dcf_report(self, 
                               ticker: str,
                               dcf_results,
                               financial_data: Dict,
                               timestamp: str) -> str:
        """Create HTML DCF report"""
        
        # Generate report content
        html_content = self._generate_html_content(ticker, dcf_results, financial_data)
        
        # Save HTML file
        filename = f"{ticker}_DCF_Report_{timestamp}.html"
        output_path = self.config.processed_data_path / filename
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        logger.info(f"HTML DCF report created: {output_path}")
        return str(output_path)
    
    def _generate_html_content(self, 
                              ticker: str,
                              dcf_results,
                              financial_data: Dict) -> str:
        """Generate HTML content for DCF report"""
        
        # Extract key data
        dcf_summary = financial_data.get('dcf_summary', {})
        current_financials = dcf_summary.get('current_financials', {})
        
        html_template = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>DCF Valuation Report - {ticker}</title>
    <style>
        body {{ 
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; 
            margin: 20px; 
            background-color: #f5f5f5;
        }}
        .container {{ 
            max-width: 1200px; 
            margin: 0 auto; 
            background-color: white; 
            padding: 30px; 
            border-radius: 10px; 
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        }}
        .header {{ 
            text-align: center; 
            color: #2c3e50; 
            border-bottom: 3px solid #3498db; 
            padding-bottom: 20px; 
            margin-bottom: 30px;
        }}
        .section {{ 
            margin: 30px 0; 
            padding: 20px; 
            border-left: 4px solid #3498db; 
            background-color: #f8f9fa;
        }}
        .highlight {{ 
            background-color: #fff3cd; 
            padding: 15px; 
            border-radius: 5px; 
            margin: 15px 0;
        }}
        .value-highlight {{ 
            font-size: 24px; 
            font-weight: bold; 
            color: #27ae60;
        }}
        .warning {{ 
            color: #e74c3c; 
            font-weight: bold;
        }}
        table {{ 
            width: 100%; 
            border-collapse: collapse; 
            margin: 20px 0;
        }}
        th, td {{ 
            border: 1px solid #ddd; 
            padding: 12px; 
            text-align: left;
        }}
        th {{ 
            background-color: #3498db; 
            color: white;
        }}
        tr:nth-child(even) {{ 
            background-color: #f2f2f2;
        }}
        .chart-container {{ 
            margin: 20px 0; 
            text-align: center;
        }}
        .footer {{ 
            text-align: center; 
            color: #7f8c8d; 
            font-size: 12px; 
            margin-top: 40px; 
            padding-top: 20px; 
            border-top: 1px solid #ddd;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>DCF Valuation Analysis</h1>
            <h2>{ticker} - {datetime.now().strftime('%B %d, %Y')}</h2>
        </div>
        
        <div class="section">
            <h3>Executive Summary</h3>
            <div class="highlight">
                <p><strong>Fair Value per Share:</strong> 
                   <span class="value-highlight">${dcf_results.value_per_share:.2f}</span></p>
                <p><strong>Current Market Price:</strong> ${dcf_results.current_price:.2f}</p>
                <p><strong>Implied Return:</strong> 
                   <span class="{'value-highlight' if dcf_results.implied_return > 0 else 'warning'}">
                   {dcf_results.implied_return:.1%}</span></p>
            </div>
        </div>
        
        <div class="section">
            <h3>Valuation Summary</h3>
            {self._create_html_valuation_table(dcf_results)}
        </div>
        
        <div class="section">
            <h3>Key Financial Metrics</h3>
            {self._create_html_financial_table(current_financials)}
        </div>
        
        <div class="section">
            <h3>DCF Assumptions</h3>
            {self._create_html_assumptions_table(dcf_results.assumptions_used)}
        </div>
        
        <div class="section">
            <h3>FCFF Projections</h3>
            {self._create_html_fcff_table(dcf_results.fcff_projections)}
        </div>
        
        <div class="section">
            <h3>Investment Recommendation</h3>
            {self._create_investment_recommendation(dcf_results)}
        </div>
        
        <div class="footer">
            <p>Generated by QuantFlow Financial Suite | 
               Report Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            <p><em>This analysis is for informational purposes only and should not be considered as investment advice.</em></p>
        </div>
    </div>
</body>
</html>
"""
        return html_template
    
    def _create_html_valuation_table(self, dcf_results) -> str:
        """Create HTML valuation summary table"""
        
        return f"""
        <table>
            <tr><th>Component</th><th>Value ($M)</th></tr>
            <tr><td>Present Value of FCFFs (Years 1-5)</td><td>${dcf_results.pv_of_fcff/1_000_000:.0f}</td></tr>
            <tr><td>Present Value of Terminal Value</td><td>${dcf_results.pv_of_terminal_value/1_000_000:.0f}</td></tr>
            <tr><td><strong>Enterprise Value</strong></td><td><strong>${dcf_results.enterprise_value/1_000_000:.0f}</strong></td></tr>
            <tr><td>Add: Net Cash</td><td>${dcf_results.net_cash/1_000_000:.0f}</td></tr>
            <tr><td><strong>Equity Value</strong></td><td><strong>${dcf_results.equity_value/1_000_000:.0f}</strong></td></tr>
            <tr><td>Shares Outstanding (M)</td><td>{dcf_results.shares_outstanding/1_000_000:.0f}</td></tr>
            <tr><td><strong>Value per Share</strong></td><td><strong>${dcf_results.value_per_share:.2f}</strong></td></tr>
        </table>
        """
    
    def _create_html_financial_table(self, current_financials: Dict) -> str:
        """Create HTML financial metrics table"""
        
        return f"""
        <table>
            <tr><th>Metric</th><th>Value ($M)</th></tr>
            <tr><td>Revenue</td><td>${current_financials.get('revenue', 0)/1_000_000:.0f}</td></tr>
            <tr><td>EBIT</td><td>${current_financials.get('ebit', 0)/1_000_000:.0f}</td></tr>
            <tr><td>Net Income</td><td>${current_financials.get('net_income', 0)/1_000_000:.0f}</td></tr>
            <tr><td>Free Cash Flow</td><td>${current_financials.get('free_cash_flow', 0)/1_000_000:.0f}</td></tr>
            <tr><td>FCFF (NOPAT Method)</td><td>${current_financials.get('fcff_nopat_method', 0)/1_000_000:.0f}</td></tr>
        </table>
        """
    
    def _create_html_assumptions_table(self, assumptions) -> str:
        """Create HTML assumptions table"""
        
        if assumptions is None:
            return "<p>No assumptions data available</p>"
        
        return f"""
        <table>
            <tr><th>Parameter</th><th>Value</th></tr>
            <tr><td>Terminal Growth Rate</td><td>{assumptions.terminal_growth_rate:.1%}</td></tr>
            <tr><td>WACC</td><td>{assumptions.wacc:.1%}</td></tr>
            <tr><td>Tax Rate</td><td>{assumptions.tax_rate:.1%}</td></tr>
            <tr><td>Gross Margin</td><td>{assumptions.gross_margin:.1%}</td></tr>
            <tr><td>R&D Margin</td><td>{assumptions.rd_margin:.1%}</td></tr>
            <tr><td>S&M Margin</td><td>{assumptions.sm_margin:.1%}</td></tr>
        </table>
        """
    
    def _create_html_fcff_table(self, fcff_projections: pd.DataFrame) -> str:
        """Create HTML FCFF projections table"""
        
        if fcff_projections is None or fcff_projections.empty:
            return "<p>No FCFF projections available</p>"
        
        # Create table headers
        html = "<table><tr><th>Year</th>"
        key_columns = ['Revenue', 'NOPAT', 'FCFF', 'PV of FCFF']
        for col in key_columns:
            if col in fcff_projections.columns:
                html += f"<th>{col} ($M)</th>"
        html += "</tr>"
        
        # Create table rows
        for year, row in fcff_projections.iterrows():
            html += f"<tr><td>{year}</td>"
            for col in key_columns:
                if col in fcff_projections.columns:
                    value = row[col] / 1_000_000 if not pd.isna(row[col]) else 0
                    html += f"<td>${value:.0f}</td>"
            html += "</tr>"
        
        html += "</table>"
        return html
    
    def _create_investment_recommendation(self, dcf_results) -> str:
        """Create investment recommendation section"""
        
        implied_return = dcf_results.implied_return
        
        if implied_return > 0.15:
            recommendation = "STRONG BUY"
            rationale = "The DCF analysis suggests the stock is significantly undervalued with high upside potential."
            color = "#27ae60"
        elif implied_return > 0.05:
            recommendation = "BUY"
            rationale = "The stock appears undervalued based on fundamental analysis."
            color = "#f39c12"
        elif implied_return > -0.05:
            recommendation = "HOLD"
            rationale = "The stock appears fairly valued at current levels."
            color = "#3498db"
        else:
            recommendation = "SELL"
            rationale = "The DCF analysis suggests the stock may be overvalued."
            color = "#e74c3c"
        
        return f"""
        <div class="highlight">
            <h4 style="color: {color};">Recommendation: {recommendation}</h4>
            <p><strong>Rationale:</strong> {rationale}</p>
            <p><strong>Key Risks:</strong></p>
            <ul>
                <li>Model assumptions may not reflect future reality</li>
                <li>Market conditions can affect valuation multiples</li>
                <li>Company-specific risks may impact cash flow projections</li>
            </ul>
        </div>
        """
    
    def export_to_excel(self, 
                       data: Union[Dict, pd.DataFrame, List[pd.DataFrame]],
                       filename: str,
                       sheet_names: Optional[List[str]] = None,
                       apply_styling: bool = True,
                       include_charts: bool = False) -> str:
        """
        Export data to Excel with professional formatting
        
        Args:
            data: Data to export (DataFrame, dict of DataFrames, or list)
            filename: Output filename
            sheet_names: Names for sheets (if multiple DataFrames)
            apply_styling: Whether to apply professional styling
            include_charts: Whether to include charts
            
        Returns:
            Path to created file
        """
        logger.info(f"Exporting data to Excel: {filename}")
        
        # Ensure output directory exists
        output_path = Path(filename)
        if not output_path.is_absolute():
            output_path = self.config.processed_data_path / output_path
        
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        try:
            with pd.ExcelWriter(str(output_path), engine='openpyxl') as writer:
                
                if isinstance(data, pd.DataFrame):
                    # Single DataFrame
                    sheet_name = sheet_names[0] if sheet_names else 'Data'
                    data.to_excel(writer, sheet_name=sheet_name, index=True)
                    
                    if apply_styling:
                        self._apply_excel_styling(writer.book[sheet_name], data)
                    
                elif isinstance(data, dict):
                    # Dictionary of DataFrames
                    for i, (key, df) in enumerate(data.items()):
                        sheet_name = sheet_names[i] if sheet_names and i < len(sheet_names) else key
                        df.to_excel(writer, sheet_name=sheet_name[:31], index=True)  # Excel sheet name limit
                        
                        if apply_styling:
                            self._apply_excel_styling(writer.book[sheet_name[:31]], df)
                
                elif isinstance(data, list):
                    # List of DataFrames
                    for i, df in enumerate(data):
                        sheet_name = sheet_names[i] if sheet_names and i < len(sheet_names) else f'Sheet{i+1}'
                        df.to_excel(writer, sheet_name=sheet_name, index=True)
                        
                        if apply_styling:
                            self._apply_excel_styling(writer.book[sheet_name], df)
                
                # Add charts if requested
                if include_charts and isinstance(data, (dict, pd.DataFrame)):
                    self._add_excel_charts(writer.book, data)
            
            # Log export
            self.export_history.append({
                'timestamp': datetime.now(),
                'filename': str(output_path),
                'format': 'Excel',
                'data_type': type(data).__name__
            })
            
            logger.info(f"Excel export completed: {output_path}")
            return str(output_path)
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            raise
    
    def _apply_excel_styling(self, worksheet, df: pd.DataFrame):
        """Apply professional styling to Excel worksheet"""
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Style headers
        for cell in worksheet[1]:  # First row (headers)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Style data cells
        for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1):
            for cell in row:
                cell.border = thin_border
                
                # Format numbers
                if isinstance(cell.value, (int, float)) and not pd.isna(cell.value):
                    if abs(cell.value) > 1000000:
                        cell.number_format = '#,##0.0,,"M"'  # Millions
                    elif abs(cell.value) > 1000:
                        cell.number_format = '#,##0.0,"K"'   # Thousands
                    elif abs(cell.value) < 1 and abs(cell.value) > 0:
                        cell.number_format = '0.0%'         # Percentages
                    else:
                        cell.number_format = '#,##0.00'     # Regular numbers
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Add conditional formatting for numerical columns
        for col_idx, col_name in enumerate(df.columns, start=2):  # Start from column B
            if df[col_name].dtype in ['int64', 'float64']:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                col_range = f"{col_letter}2:{col_letter}{len(df)+1}"
                
                # Color scale: red for low values, green for high values
                color_scale = ColorScaleRule(
                    start_type='min', start_color='FF6B6B',
                    end_type='max', end_color='4ECDC4'
                )
                worksheet.conditional_formatting.add(col_range, color_scale)
    
    def _add_excel_charts(self, workbook, data: Union[Dict, pd.DataFrame]):
        """Add charts to Excel workbook"""
        
        try:
            if isinstance(data, pd.DataFrame):
                self._create_chart_for_dataframe(workbook, data, 'Data')
            elif isinstance(data, dict):
                for sheet_name, df in data.items():
                    if isinstance(df, pd.DataFrame) and not df.empty:
                        self._create_chart_for_dataframe(workbook, df, sheet_name[:31])
        except Exception as e:
            logger.warning(f"Could not create charts: {e}")
    
    def _create_chart_for_dataframe(self, workbook, df: pd.DataFrame, sheet_name: str):
        """Create a chart for a specific DataFrame"""
        
        worksheet = workbook[sheet_name]
        
        # Find numerical columns for charting
        numerical_cols = df.select_dtypes(include=[np.number]).columns[:5]  # Limit to 5 series
        
        if len(numerical_cols) > 0 and len(df) > 1:
            # Create line chart
            chart = LineChart()
            chart.title = f"{sheet_name} Trends"
            chart.style = 13
            chart.y_axis.title = 'Values'
            chart.x_axis.title = 'Period'
            
            # Add data series
            for col_idx, col_name in enumerate(numerical_cols):
                col_letter = openpyxl.utils.get_column_letter(col_idx + 2)  # Skip index column
                data_range = Reference(worksheet, 
                                     min_col=col_idx + 2, 
                                     min_row=1, 
                                     max_row=len(df) + 1)
                chart.add_data(data_range, titles_from_data=True)
            
            # Position chart
            chart.width = 15
            chart.height = 10
            chart_cell = f"A{len(df) + 5}"  # Position below data
            worksheet.add_chart(chart, chart_cell)
    
    def export_to_csv(self, 
                     data: Union[pd.DataFrame, Dict[str, pd.DataFrame]],
                     filename: str,
                     **kwargs) -> str:
        """
        Export data to CSV format
        
        Args:
            data: DataFrame or dict of DataFrames
            filename: Output filename
            **kwargs: Additional arguments for to_csv()
            
        Returns:
            Path to created file(s)
        """
        logger.info(f"Exporting data to CSV: {filename}")
        
        output_path = Path(filename)
        if not output_path.is_absolute():
            output_path = self.config.processed_data_path / output_path
        
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        if isinstance(data, pd.DataFrame):
            # Single DataFrame
            data.to_csv(output_path, **kwargs)
            return str(output_path)
        
        elif isinstance(data, dict):
            # Multiple DataFrames - create separate files
            exported_files = []
            for name, df in data.items():
                file_path = output_path.parent / f"{output_path.stem}_{name}.csv"
                df.to_csv(file_path, **kwargs)
                exported_files.append(str(file_path))
            
            return exported_files[0] if len(exported_files) == 1 else exported_files
    
    def export_to_json(self, 
                      data: Union[Dict, pd.DataFrame, List],
                      filename: str,
                      orient: str = 'records',
                      indent: int = 2) -> str:
        """
        Export data to JSON format
        
        Args:
            data: Data to export
            filename: Output filename
            orient: DataFrame orientation for JSON
            indent: JSON indentation
            
        Returns:
            Path to created file
        """
        logger.info(f"Exporting data to JSON: {filename}")
        
        output_path = Path(filename)
        if not output_path.is_absolute():
            output_path = self.config.processed_data_path / output_path
        
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Convert data to JSON-serializable format
        if isinstance(data, pd.DataFrame):
            json_data = data.to_dict(orient=orient)
        elif isinstance(data, dict):
            json_data = {}
            for key, value in data.items():
                if isinstance(value, pd.DataFrame):
                    json_data[key] = value.to_dict(orient=orient)
                else:
                    json_data[key] = self._make_json_serializable(value)
        else:
            json_data = self._make_json_serializable(data)
        
        with open(output_path, 'w') as f:
            json.dump(json_data, f, indent=indent, default=str)
        
        return str(output_path)
    
    def export_dcf_model_to_excel(self, 
                                 dcf_results,
                                 financial_data: Dict,
                                 filename: str) -> str:
        """
        Export complete DCF model to Excel with Adobe-style formatting
        
        Args:
            dcf_results: DCF calculation results
            financial_data: Complete financial data
            filename: Output filename
            
        Returns:
            Path to created Excel file
        """
        logger.info(f"Exporting DCF model to Excel: {filename}")
        
        output_path = Path(filename)
        if not output_path.is_absolute():
            output_path = self.config.processed_data_path / output_path
        
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Prepare data for export
        export_data = {
            'DCF Summary': self._create_dcf_summary_table(dcf_results),
            'FCFF Projections': dcf_results.fcff_projections if dcf_results.fcff_projections is not None else pd.DataFrame(),
            'Historical Financials': self._create_historical_summary(financial_data),
            'Assumptions': self._create_assumptions_table(dcf_results.assumptions_used),
            'Sensitivity Analysis': dcf_results.sensitivity_table if dcf_results.sensitivity_table is not None else pd.DataFrame()
        }
        
        # Export with custom styling
        excel_path = self.export_to_excel(
            export_data, 
            str(output_path),
            apply_styling=True,
            include_charts=True
        )
        
        # Add DCF-specific formatting
        self._apply_dcf_specific_formatting(excel_path, dcf_results)
        
        return excel_path
    
    def _create_dcf_summary_table(self, dcf_results) -> pd.DataFrame:
        """Create DCF summary table"""
        
        summary_data = {
            'Component': [
                'Present Value of FCFFs (Years 1-5)',
                'Present Value of Terminal Value',
                'Enterprise Value',
                'Add: Net Cash',
                'Equity Value',
                'Shares Outstanding (millions)',
                'Value per Share',
                'Current Stock Price',
                'Implied Return'
            ],
            'Value ($M except per share)': [
                dcf_results.pv_of_fcff / 1_000_000,
                dcf_results.pv_of_terminal_value / 1_000_000,
                dcf_results.enterprise_value / 1_000_000,
                dcf_results.net_cash / 1_000_000,
                dcf_results.equity_value / 1_000_000,
                dcf_results.shares_outstanding / 1_000_000,
                dcf_results.value_per_share,
                dcf_results.current_price,
                dcf_results.implied_return
            ]
        }
        
        return pd.DataFrame(summary_data)
    
    def _create_historical_summary(self, financial_data: Dict) -> pd.DataFrame:
        """Create historical financial summary"""
        
        try:
            income_df = financial_data.get('income_statement', pd.DataFrame())
            cashflow_df = financial_data.get('cashflow_statement', pd.DataFrame())
            
            if income_df.empty or cashflow_df.empty:
                return pd.DataFrame()
            
            # Get common years
            common_years = income_df.index.intersection(cashflow_df.index)[:5]  # Last 5 years
            
            historical_data = []
            for year in common_years:
                historical_data.append({
                    'Year': year.strftime('%Y'),
                    'Revenue ($M)': income_df.loc[year].get('Total Revenue', 0) / 1_000_000,
                    'EBIT ($M)': income_df.loc[year].get('Operating Income (EBIT)', 0) / 1_000_000,
                    'Net Income ($M)': income_df.loc[year].get('Net Income', 0) / 1_000_000,
                    'Free Cash Flow ($M)': cashflow_df.loc[year].get('Free Cash Flow', 0) / 1_000_000,
                    'FCFF ($M)': cashflow_df.loc[year].get('FCFF (NOPAT Method)', 0) / 1_000_000
                })
            
            return pd.DataFrame(historical_data)
            
        except Exception as e:
            logger.warning(f"Could not create historical summary: {e}")
            return pd.DataFrame()
    
    def _create_assumptions_table(self, assumptions) -> pd.DataFrame:
        """Create assumptions table"""
        
        if assumptions is None:
            return pd.DataFrame()
        
        assumptions_data = {
            'Parameter': [
                'Projection Years',
                'Terminal Growth Rate',
                'Tax Rate',
                'WACC',
                'Gross Margin',
                'R&D as % of Revenue',
                'S&M as % of Revenue',
                'G&A as % of Revenue',
                'D&A as % of Revenue',
                'CapEx as % of Revenue',
                'NWC Growth Factor'
            ],
            'Value': [
                assumptions.projection_years,
                f"{assumptions.terminal_growth_rate:.1%}",
                f"{assumptions.tax_rate:.1%}",
                f"{assumptions.wacc:.1%}",
                f"{assumptions.gross_margin:.1%}",
                f"{assumptions.rd_margin:.1%}",
                f"{assumptions.sm_margin:.1%}",
                f"{assumptions.ga_margin:.1%}",
                f"{assumptions.da_margin:.1%}",
                f"{assumptions.capex_margin:.1%}",
                f"{assumptions.nwc_growth_factor:.1%}"
            ],
            'Notes': [
                'Years of explicit projections',
                'Long-term growth rate beyond projection period',
                'Corporate tax rate for NOPAT calculation',
                'Weighted Average Cost of Capital',
                'Gross profit margin assumption',
                'Research & Development expense ratio',
                'Sales & Marketing expense ratio',
                'General & Administrative expense ratio',
                'Depreciation & Amortization ratio',
                'Capital Expenditure ratio',
                'Working capital change factor'
            ]
        }
        
        return pd.DataFrame(assumptions_data)
    
    def _apply_dcf_specific_formatting(self, excel_path: str, dcf_results):
        """Apply DCF-specific formatting to Excel file"""
        
        try:
            workbook = openpyxl.load_workbook(excel_path)
            
            # Format DCF Summary sheet
            if 'DCF Summary' in workbook.sheetnames:
                summary_sheet = workbook['DCF Summary']
                
                # Highlight key results
                key_rows = [9, 10, 11]  # Value per share, current price, implied return
                for row in key_rows:
                    for cell in summary_sheet[row]:
                        if cell.value is not None:
                            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                            cell.font = Font(bold=True)
            
            # Format Sensitivity Analysis sheet
            if 'Sensitivity Analysis' in workbook.sheetnames and dcf_results.sensitivity_table is not None:
                sens_sheet = workbook['Sensitivity Analysis']
                
                # Add color scaling to sensitivity table
                max_row = len(dcf_results.sensitivity_table) + 1
                max_col = len(dcf_results.sensitivity_table.columns) + 1
                
                sens_range = f"B2:{openpyxl.utils.get_column_letter(max_col)}{max_row}"
                color_scale = ColorScaleRule(
                    start_type='percentile', start_value=10, start_color='FF6B6B',
                    mid_type='percentile', mid_value=50, mid_color='FFEB9C',
                    end_type='percentile', end_value=90, end_color='4ECDC4'
                )
                sens_sheet.conditional_formatting.add(sens_range, color_scale)
            
            workbook.save(excel_path)
            
        except Exception as e:
            logger.warning(f"Could not apply DCF-specific formatting: {e}")
    
    def _make_json_serializable(self, obj):
        """Convert objects to JSON-serializable format"""
        
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        elif isinstance(obj, np.integer):
            return int(obj)
        elif isinstance(obj, np.floating):
            return float(obj)
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        elif hasattr(obj, '__dict__'):
            return {k: self._make_json_serializable(v) for k, v in obj.__dict__.items()}
        else:
            return obj
    
class DataLoader:
    """
    Comprehensive data loading utility for QuantFlow
    Supports multiple file formats and data validation
    """
    
    def __init__(self, config=None):
        self.config = config or get_config()
    
    def load_from_excel(self, 
                       filepath: str,
                       sheet_name: Optional[Union[str, List[str]]] = None,
                       **kwargs) -> Union[pd.DataFrame, Dict[str, pd.DataFrame]]:
        """
        Load data from Excel file
        
        Args:
            filepath: Path to Excel file
            sheet_name: Sheet name(s) to load
            **kwargs: Additional arguments for read_excel()
            
        Returns:
            DataFrame or dict of DataFrames
        """
        logger.info(f"Loading data from Excel: {filepath}")
        
        try:
            if sheet_name is None:
                # Load all sheets
                data = pd.read_excel(filepath, sheet_name=None, **kwargs)
                return data
            else:
                # Load specific sheet(s)
                data = pd.read_excel(filepath, sheet_name=sheet_name, **kwargs)
                return data
                
        except Exception as e:
            logger.error(f"Error loading Excel file: {e}")
            raise
    
    def load_from_csv(self, 
                     filepath: str,
                     **kwargs) -> pd.DataFrame:
        """
        Load data from CSV file
        
        Args:
            filepath: Path to CSV file
            **kwargs: Additional arguments for read_csv()
            
        Returns:
            DataFrame
        """
        logger.info(f"Loading data from CSV: {filepath}")
        
        try:
            data = pd.read_csv(filepath, **kwargs)
            return data
        except Exception as e:
            logger.error(f"Error loading CSV file: {e}")
            raise
    
    def load_from_json(self, 
                      filepath: str,
                      **kwargs) -> Union[Dict, pd.DataFrame]:
        """
        Load data from JSON file
        
        Args:
            filepath: Path to JSON file
            **kwargs: Additional arguments for read_json()
            
        Returns:
            Dictionary or DataFrame
        """
        logger.info(f"Loading data from JSON: {filepath}")
        
        try:
            with open(filepath, 'r') as f:
                data = json.load(f)
            return data
        except Exception as e:
            logger.error(f"Error loading JSON file: {e}")
            raise
    
    def load_configuration(self, 
                          config_path: str) -> Dict:
        """
        Load configuration from YAML or JSON file
        
        Args:
            config_path: Path to configuration file
            
        Returns:
            Configuration dictionary
        """
        logger.info(f"Loading configuration from: {config_path}")
        
        config_path = Path(config_path)
        
        try:
            if config_path.suffix.lower() in ['.yaml', '.yml']:
                with open(config_path, 'r') as f:
                    config = yaml.safe_load(f)
            elif config_path.suffix.lower() == '.json':
                with open(config_path, 'r') as f:
                    config = json.load(f)
            else:
                raise ValueError(f"Unsupported configuration file format: {config_path.suffix}")
            
            return config
            
        except Exception as e:
            logger.error(f"Error loading configuration: {e}")
            raise


class ReportGenerator:
    """
    Professional report generator for QuantFlow analysis
    Creates formatted reports with charts and tables
    """

    def __init__(self, config=None):
        self.config = config or get_config()

    def create_dcf_report(self, ticker, dcf_results, financial_data, format="excel"):
        from pathlib import Path
        import pandas as pd

        # Use existing 'data/processed/' folder
        output_dir = Path("data/processed")
        file_path = output_dir / f"{ticker}_dcf_report.xlsx"

        with pd.ExcelWriter(file_path, engine="xlsxwriter") as writer:
            summary_df = pd.DataFrame({
                "Metric": ["Value Per Share", "Current Price", "Implied Return", "Enterprise Value", "Equity Value"],
                "Value": [
                    dcf_results.value_per_share,
                    dcf_results.current_price,
                    dcf_results.implied_return,
                    dcf_results.enterprise_value,
                    dcf_results.equity_value
                ]
            })
            summary_df.to_excel(writer, sheet_name="Summary", index=False)

            if hasattr(dcf_results, "fcff_projections") and dcf_results.fcff_projections is not None:
                dcf_results.fcff_projections.to_excel(writer, sheet_name="Projections")

            if hasattr(dcf_results, "sensitivity_table") and dcf_results.sensitivity_table is not None:
                dcf_results.sensitivity_table.to_excel(writer, sheet_name="Sensitivity")

        return str(file_path)


# Convenience functions
def export_financial_data(financial_data: Dict, 
                         ticker: str,
                         format: str = 'excel') -> str:
    """
    Export comprehensive financial data
    
    Usage:
        path = export_financial_data(data, "ADBE", "excel")
    """
    exporter = DataExporter()
    
    if format == 'excel':
        export_data = {
            'Income Statement': financial_data.get('income_statement', pd.DataFrame()),
            'Balance Sheet': financial_data.get('balance_sheet', pd.DataFrame()),
            'Cash Flow': financial_data.get('cashflow_statement', pd.DataFrame()),
            'Key Ratios': financial_data.get('comprehensive_ratios', pd.DataFrame())
        }
        return exporter.export_to_excel(export_data, f"{ticker}_Financial_Data.xlsx")
    
    elif format == 'csv':
        # Export each statement as separate CSV
        results = []
        for name, df in financial_data.items():
            if isinstance(df, pd.DataFrame) and not df.empty:
                filename = f"{ticker}_{name}.csv"
                path = exporter.export_to_csv(df, filename)
                results.append(path)
        return results
    
    elif format == 'json':
        return exporter.export_to_json(financial_data, f"{ticker}_Financial_Data.json")


def create_dcf_report(ticker: str, 
                     dcf_results,
                     financial_data: Dict,
                     format: str = 'excel') -> str:
    """
    Create comprehensive DCF report
    
    Usage:
        report_path = create_dcf_report("ADBE", dcf_results, financial_data, "html")
    """
    report_generator = ReportGenerator()
    return report_generator.create_dcf_report(ticker, dcf_results, financial_data, format)


def load_financial_template(template_name: str) -> pd.DataFrame:
    """
    Load financial analysis template
    
    Usage:
        template = load_financial_template("dcf_template.xlsx")
    """
    loader = DataLoader()
    template_path = Path(__file__).parent.parent / "templates" / template_name
    
    if template_path.exists():
        return loader.load_from_excel(str(template_path))
    else:
        logger.warning(f"Template not found: {template_path}")
        return pd.DataFrame()

def batch_export_companies(tickers: List[str], 
                          format: str = 'excel',
                          include_dcf: bool = True) -> List[str]:
    """
    Export financial data for multiple companies
    
    Usage:
        paths = batch_export_companies(["ADBE", "MSFT", "GOOGL"], "excel", True)
    """
    from ..fetchers import fetch_comprehensive_data
    from ..models.dcf_calculator import quick_dcf_valuation
    
    exported_files = []
    
    for ticker in tickers:
        try:
            logger.info(f"Processing {ticker}")
            
            # Fetch financial data
            financial_data = fetch_comprehensive_data(ticker)
            
            # Export financial data
            financial_path = export_financial_data(financial_data, ticker, format)
            exported_files.append(financial_path)
            
            # Export DCF analysis if requested
            if include_dcf:
                try:
                    dcf_results = quick_dcf_valuation(ticker)
                    dcf_path = create_dcf_report(ticker, dcf_results, financial_data, format)
                    exported_files.append(dcf_path)
                except Exception as e:
                    logger.warning(f"Could not create DCF report for {ticker}: {e}")
            
        except Exception as e:
            logger.error(f"Error processing {ticker}: {e}")
    
    return exported_files


if __name__ == "__main__":
    # Example usage
    print("📁 Testing I/O Utilities")
    
    try:
        # Test data export
        sample_data = pd.DataFrame({
            'Year': [2023, 2022, 2021],
            'Revenue': [100000, 90000, 80000],
            'Net Income': [15000, 12000, 10000]
        })
        
        exporter = DataExporter()
        
        # Test Excel export
        excel_path = exporter.export_to_excel(sample_data, "test_export.xlsx")
        print(f"✅ Excel export test: {excel_path}")
        
        # Test CSV export
        csv_path = exporter.export_to_csv(sample_data, "test_export.csv")
        print(f"✅ CSV export test: {csv_path}")
        
        # Test JSON export
        json_path = exporter.export_to_json(sample_data, "test_export.json")
        print(f"✅ JSON export test: {json_path}")
        
        # Test data loading
        loader = DataLoader()
        loaded_data = loader.load_from_csv(csv_path)
        print(f"✅ Data loading test: {len(loaded_data)} rows loaded")
        
        # Test with real financial data
        print(f"\n🧪 Testing with real financial data...")
        from ..fetchers import fetch_comprehensive_data
        
        financial_data = fetch_comprehensive_data("ADBE")
        financial_export_path = export_financial_data(financial_data, "ADBE", "excel")
        print(f"✅ Financial data export: {financial_export_path}")
        
        print(f"\n📊 I/O Utilities testing complete!")
        
    except Exception as e:
        print(f"❌ Error in I/O utilities testing: {e}")
        import traceback
        traceback.print_exc()